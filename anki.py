import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl as opyxl
from openpyxl import Workbook
from random import randint
import pandas as pd
import os

class GUI:
    def __init__(self):
        self.row = []
        self.directory = None

        self.root = tk.Tk()
        self.root.geometry("330x380")
        self.root.title("Auto Anki")
        self.root.resizable(False, False)
        
        self.root.iconbitmap("auto_anki.ico")

        self.mainframe = tk.Frame(self.root, background="#2A324B")
        self.mainframe.pack(fill="both", expand=1)

        # Create and place widgets
        self.create_widgets()

        # Show the window
        self.root.mainloop()

    def create_widgets(self):
        # Kanji
        ttk.Label(self.mainframe, text="Kanji", background="#2A324B", font=("helvetica", 24), foreground="#E1E5EE").grid(row=0, column=0, padx=15, pady=(15,0), sticky="NWES")
        self.set_kanji_field = ttk.Entry(self.mainframe, style='EntryBackground.TEntry')
        self.set_kanji_field.grid(row=1, column=0, padx=15, pady=5, sticky="NWES")

        # Hiragana
        ttk.Label(self.mainframe, text="Hiragana", background="#2A324B", font=("helvetica", 24), foreground="#E1E5EE").grid(row=2, column=0, padx=15, pady=(15,0), sticky="NWES")
        self.set_hiragana_field_1 = ttk.Entry(self.mainframe)
        self.set_hiragana_field_1.grid(row=3, column=0, padx=15, pady=5, sticky="NWES")
        self.set_hiragana_field_2 = ttk.Entry(self.mainframe)
        self.set_hiragana_field_2.grid(row=4, column=0, padx=15, pady=5, sticky="NWES")

        # English
        ttk.Label(self.mainframe, text="English", background="#2A324B", font=("helvetica", 24), foreground="#E1E5EE").grid(row=2, column=1, padx=15, pady=(15,0), sticky="NWES")
        self.set_english_field_1 = ttk.Entry(self.mainframe)
        self.set_english_field_1.grid(row=3, column=1, padx=15, pady=5, sticky="NWES")
        self.set_english_field_2 = ttk.Entry(self.mainframe)
        self.set_english_field_2.grid(row=4, column=1, padx=15, pady=5, sticky="NWES")

        # Checkbox
        self.checkbox = tk.IntVar()
        self.checkbox.set(1)
        check = ttk.Checkbutton(self.mainframe, text="Two Meanings", variable=self.checkbox, onvalue=2, offvalue=1, style='#2A324B.TCheckbutton', takefocus=False)
        check.configure(style='#2A324B.TCheckbutton')
        style = ttk.Style()
        style.configure('#2A324B.TCheckbutton', background="#2A324B", foreground="#E1E5EE", font=("helvetica", 12))
        check.grid(row=1, column=1, padx=15, pady=5, sticky="NWES")

        # Directory button
        self.directory_button = ttk.Button(self.mainframe, text="Set Excel Sheet", takefocus=False, command=self.find_directory)
        self.directory_button.grid(row=9, column=0, padx=15, pady=(40,0), sticky="NWES")

        # Insert button
        self.set_button = tk.Button(self.mainframe, text="Insert", takefocus=False, command=self.insert_to_excel, bg='#5B9BCF', relief="flat", activebackground="lightblue", font=("helvetica", 20), cursor="hand2")
        self.set_button.grid(row=9, column=1, pady=(20,0), sticky="NWES")
        self.set_button.bind("<Enter>", self.on_enter_setButton)
        self.set_button.bind("<Leave>", self.on_leave_setButton)

        # Save button
        self.save_button = tk.Button(self.mainframe, text="Save CSV", takefocus=False, command=self.save_csv, bg='#3CBF5B', relief="flat", activebackground="lightgreen", font=("helvetica", 20), cursor="hand2")
        self.save_button.grid(row=10, column=0, columnspan=2, pady=(20,0), padx=(15,0), sticky="EW")
        self.save_button.bind("<Enter>", self.on_enter_saveButton)
        self.save_button.bind("<Leave>", self.on_leave_saveButton)

    def find_directory(self):
        self.directory = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        print(self.directory)

    def insert_to_excel(self):
        if not self.directory:
            messagebox.showerror("Excel file not found", "Error: No directory selected. Please select an Excel file first.")
            return

        try:
            if self.checkbox.get() == 1:
                self.row = [
                    self.checkbox.get(),
                    self.set_kanji_field.get(),
                    self.set_hiragana_field_1.get(),
                    self.set_english_field_1.get()
                ]
            elif self.checkbox.get() == 2:
                self.row = [
                    self.checkbox.get(),
                    self.set_kanji_field.get(),
                    self.set_hiragana_field_1.get(),
                    self.set_english_field_1.get(),
                    self.set_hiragana_field_2.get(),
                    self.set_english_field_2.get()
                ]

            exporter = Anki(self.directory, self.row)
            exporter.export()

            self.reset_fields()


        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")
    
    def reset_fields(self):
        self.set_kanji_field.delete(0, tk.END)
        self.set_hiragana_field_1.delete(0, tk.END)
        self.set_hiragana_field_2.delete(0, tk.END)
        self.set_english_field_1.delete(0, tk.END)
        self.set_english_field_2.delete(0, tk.END)
    
    def save_csv(self):
        try:
            directory_path = os.path.dirname(self.directory)
            filename = os.path.splitext(os.path.basename(self.directory))[0] + " CSV.csv"
            filepath = os.path.join(directory_path, filename)
            df = pd.read_excel(self.directory)
            df.to_csv(filepath, index=False, encoding='utf-8')

            messagebox.showinfo("Success", "CSV file saved successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

    def on_enter_setButton(self, event):
        self.set_button['background'] = "lightblue"

    def on_leave_setButton(self, event):
        self.set_button['background'] = "#5B9BCF"

    def on_enter_saveButton(self, event):
        self.save_button['background'] = "lightgreen"

    def on_leave_saveButton(self, event):
        self.save_button['background'] = "#3CBF5B"

class Anki:
    def __init__(self, excel_file: str, row):
        self.excel_file = excel_file
        self.workbook = opyxl.load_workbook(excel_file)
        self.worksheet = self.workbook.active
        self.row = row

    def front(self) -> str:
        frontcard = f"<H1>{self.row[1]}</H1>"
        return frontcard

    def back(self) -> str:
        amount = self.row[0]
        jap1 = self.row[2]
        en1 = self.row[3]

        if amount == 1:
            audio = self.row[3]
            image = self.row[3]
            random_integer = randint(0, 1)
            if random_integer == 0:
                #blue and green
                html = f"""<div class="main"><div class="text"><div><span class="flavor1">{jap1}</span></div><div><span class="flavor1">{en1}</span></div><div>[sound:{audio}.mp3]</div></div><img src="{image}.webp"></div><style>div{{font-size:50px;}}.main{{display:flex;flex-direction:row;align-items:center;justify-content:space-evenly;}}.text{{display:flex;flex-direction:column;justify-content:center;align-items:center;gap:18px;font-weight:600;}}.flavor1{{color:rgb(155,255,178);background-color:rgb(50,85,200);padding:5px 10px;border-radius:10px;}}.flavor2{{color:rgb(255,170,255);background-color:rgb(85,0,255);padding:5px 10px;border-radius:10px;}}</style>"""
            elif random_integer == 1:
                #pink and purple
                html = f"""<div class="main"><div class="text"><div><span class="flavor2">{jap1}</span></div><div><span class="flavor2">{en1}</span></div><div>[sound:{audio}.mp3]</div></div><img src="{image}.webp"></div><style>div{{font-size:50px;}}.main{{display:flex;flex-direction:row;align-items:center;justify-content:space-evenly;}}.text{{display:flex;flex-direction:column;justify-content:center;align-items:center;gap:18px;font-weight:600;}}.flavor1{{color:rgb(155,255,178);background-color:rgb(50,85,200);padding:5p x10px;border-radius:10px;}}.flavor2{{color:rgb(255,170,255);background-color:rgb(85,0,255);padding:5px 10px;border-radius:10px;}}</style>"""
        elif amount == 2:
            jap2 = self.row[4]
            en2 = self.row[5]
            audio1 = self.row[3]
            audio2 = self.row[5]
            image = self.row[3]
            html = f"""<div class="main"><div class="text"><div><span class="flavor1">{jap1}</span>|<span class="flavor2">{jap2}</span></div><div><span class="flavor1">{en1}</span>|<span class="flavor2">{en2}</span></div><div>[sound:{audio1}.mp3][sound:{audio2}.mp3]</div></div><img src="{image}.png"></d x;border-radius:10px;}}</style>"""

        return html

    def export(self):
        for row in self.worksheet.iter_rows(min_col=1, max_col=1):
            cell = row[0]
        self.worksheet.cell(row=cell.row+1, column=1, value=self.front())
        self.worksheet.cell(row=cell.row+1, column=2, value=self.back())

        self.workbook.save(self.excel_file)


if __name__ == "__main__":
    GUI()